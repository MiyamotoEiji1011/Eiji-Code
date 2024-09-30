import discord
from discord.ext import commands
import pandas as pd
import openpyxl
import config

intents = discord.Intents.default()
intents.message_content = True

bot = commands.Bot(command_prefix='!todo ', intents=intents)

# グローバル変数で現在のシート名を保持
current_sheet_name = None

def add_sheet(file_name, sheet_name):
    try:
        wb = openpyxl.load_workbook(file_name)
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(title=sheet_name)
            ws = wb[sheet_name]
            ws.append(["No", "State", "Task", "Content", "data", "user"])
            wb.save(file_name)
            print(f'Sheet "{sheet_name}" added successfully.')
        else:
            print(f'Sheet "{sheet_name}" already exists.')
    except Exception as e:
        print(f'Error: {str(e)}')

def delete_sheet(file_name, sheet_name):
    try:
        wb = openpyxl.load_workbook(file_name)
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
            wb.save(file_name)
            print(f'Sheet "{sheet_name}" deleted successfully.')
        else:
            print(f'Sheet "{sheet_name}" not found.')
    except Exception as e:
        print(f'Error: {str(e)}')

@bot.event
async def on_ready():
    print(f'Logged in as {bot.user.name}')
    
    # チャンネルIDを指定
    channel_id = config.ALLOWED_CHANNEL_ID
    channel = bot.get_channel(channel_id)
    
    if channel:
        embed = discord.Embed(
            title='🚀 Tasky is Ready!',
            description='🌟 Hello everyone! Tasky is now online and ready to help you manage your tasks efficiently. \n\n💡 Use `!todo` commands to create, view, and manage your to-do lists with ease. \n\nLet\'s get organized and make things happen! 🚀✨',
            color=0x00ff00
        )
        await channel.send(embed=embed)
    else:
        print('Channel not found!')



async def send_embed(ctx, title, description, color=0xffc0cb):
    embed = discord.Embed(title=title, description=description, color=color)
    await ctx.send(embed=embed)

@bot.command()
async def set(ctx, sheet_name: str):
    global current_sheet_name
    xls = pd.ExcelFile('todo_list.xlsx', engine='openpyxl')
    if sheet_name in xls.sheet_names:
        current_sheet_name = sheet_name
        await send_embed(ctx, '', f'Sheet set to {sheet_name}')
        await show(ctx)
    else:
        await send_embed(ctx, '', f'Sheet "{sheet_name}" not found.', color=0xff0000)


@bot.command()
async def show(ctx):
    if current_sheet_name is None:
        await send_embed(ctx, '', 'No sheet is set. Use `!todo set "sheet_name"` first.', color=0xff0000)
        return

    df = pd.read_excel('todo_list.xlsx', sheet_name=current_sheet_name)
    try:
        if pd.api.types.is_numeric_dtype(df['data']):
            df['data'] = pd.to_datetime(df['data'], unit='d', origin='1899-12-30')
        else:
            df['data'] = pd.to_datetime(df['data'])
    except Exception as e:
        await send_embed(ctx, '', f'Error while converting dates: {str(e)}', color=0xff0000)
        return

    for i in range(0, len(df), 10):
        description = "**No**/　**State**　[　**Task**　]　<-　**Content**　|　#**Date**　(**＠User**)\n" + "-"*60 + "\n"
        for _, row in df.iloc[i:i+10].iterrows():
            date_str = row['data'].strftime('%Y/%m/%d')
            state = ":white_check_mark:" if row['State'] else ":x:"
            task_row = (f"**No.{row['No']} / {state} [ {row['Task']} ]**　<- {row['Content']} | "
                        f"#{date_str} ( {row['user']} )")
            description += f"{task_row}\n"
        
        await send_embed(ctx, f"To-Do! - {current_sheet_name}", description , color=0x00ff00)

@bot.command()
async def add(ctx, task: str, content: str, date: str, user: str):
    if current_sheet_name is None:
        await send_embed(ctx, '', 'No sheet is set. Use `!todo set "sheet_name"` first.', color=0xff0000)
        return

    try:
        # Excelファイルを読み込む
        wb = openpyxl.load_workbook('todo_list.xlsx')
        ws = wb[current_sheet_name]
        data = ws.values
        columns = next(data)
        df = pd.DataFrame(data, columns=columns)

        # 日付のフォーマットを整える
        formatted_date = pd.to_datetime(date).strftime('%Y/%m/%d')

        # 新しいタスクを作成
        new_task = pd.DataFrame([{
            "No": len(df) + 1,
            "State": False,
            "Task": task,
            "Content": content,
            "data": formatted_date,
            "user": user
        }])

        # 新しいタスクを既存のDataFrameに追加
        df = pd.concat([df, new_task], ignore_index=True)

        # Excelファイルに書き込む
        with pd.ExcelWriter('todo_list.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=current_sheet_name, index=False)
        
        await send_embed(ctx, '', f'Task "{task}" added to sheet "{current_sheet_name}".')
        await show(ctx)
    except Exception as e:
        await send_embed(ctx, '', f'Error: {str(e)}', color=0xff0000)


@bot.command()
async def create(ctx, sheet_name: str):
    try:
        add_sheet('todo_list.xlsx', sheet_name)
        await send_embed(ctx, '', f'New sheet "{sheet_name}" has been created.')
    except Exception as e:
        await send_embed(ctx, '', f'Error: {str(e)}', color=0xff0000)

@bot.command()
async def delete(ctx, sheet_name: str):
    try:
        delete_sheet('todo_list.xlsx', sheet_name)
        global current_sheet_name
        current_sheet_name = None
        await send_embed(ctx, '', f'Sheet "{sheet_name}" has been deleted.')
    except Exception as e:
        await send_embed(ctx, '', f'Error: {str(e)}', color=0xff0000)

@bot.command()
async def done(ctx, no: int):
    if current_sheet_name is None:
        await send_embed(ctx, '', 'No sheet is set. Use `!todo set "sheet_name"` first.', color=0xff0000)
        return

    try:
        df = pd.read_excel('todo_list.xlsx', sheet_name=current_sheet_name)
        if 'No' not in df.columns:
            await send_embed(ctx, '', 'No column not found.', color=0xff0000)
            return

        if no in df['No'].values:
            df.loc[df['No'] == no, 'State'] = True
            with pd.ExcelWriter('todo_list.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=current_sheet_name, index=False)
            await send_embed(ctx, '', f'Task No.{no} marked as done!')
            await show(ctx)
        else:
            await send_embed(ctx, '', f'Task No.{no} not found in the current sheet.', color=0xff0000)
    except Exception as e:
        await send_embed(ctx, '', f'Error: {str(e)}', color=0xff0000)

@bot.command()
async def remove(ctx, no: int):
    if current_sheet_name is None:
        await send_embed(ctx, '', 'No sheet is set. Use `!todo set "sheet_name"` first.', color=0xff0000)
        return

    try:
        df = pd.read_excel('todo_list.xlsx', sheet_name=current_sheet_name)
        if 'No' not in df.columns:
            await send_embed(ctx, '', 'No column not found.', color=0xff0000)
            return

        if no in df['No'].values:
            df = df[df['No'] != no]
            
            # No列の再整理
            df.reset_index(drop=True, inplace=True)
            df['No'] = df.index + 1

            with pd.ExcelWriter('todo_list.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=current_sheet_name, index=False)
            
            await send_embed(ctx, '', f'Task No.{no} removed from the sheet and task numbers have been renumbered.')
            await show(ctx)
        else:
            await send_embed(ctx, '', f'Task No.{no} not found in the current sheet.', color=0xff0000)
    except Exception as e:
        await send_embed(ctx, '', f'Error: {str(e)}', color=0xff0000)

@bot.command()
async def user(ctx, user: str):
    if current_sheet_name is None:
        await send_embed(ctx, '', 'No sheet is set. Use `!todo set "sheet_name"` first.', color=0xff0000)
        return

    try:
        df = pd.read_excel('todo_list.xlsx', sheet_name=current_sheet_name)

        if 'user' not in df.columns:
            await send_embed(ctx, '', 'User column not found.', color=0xff0000)
            return

        # ユーザー名が含まれるタスクをフィルタリング
        filtered_df = df[df['user'].str.contains(user, case=False, na=False)]

        if filtered_df.empty:
            await send_embed(ctx, '', f'No tasks found for user containing "{user}".')
            return

        # data列をdatetime型に変換
        if not pd.api.types.is_datetime64_any_dtype(filtered_df['data']):
            filtered_df['data'] = pd.to_datetime(filtered_df['data'], errors='coerce')

        # タスクの表示
        for i in range(0, len(filtered_df), 10):
            description = "**No**/　**State**　[　**Task**　]　<-　**Content**　|　#**Date**　(**＠User**)\n" + "-"*60 + "\n"
            for _, row in filtered_df.iloc[i:i+10].iterrows():
                date_str = row['data'].strftime('%Y/%m/%d') if pd.notna(row['data']) else 'N/A'
                state = ":white_check_mark:" if row['State'] else ":x:"
                task_row = (f"**No.{row['No']} / {state} [ {row['Task']} ]**　<- {row['Content']} | "
                            f"#{date_str} ( {row['user']} )")
                description += f"{task_row}\n"

            await send_embed(ctx, f"To-Do! - {current_sheet_name} (User: {user})", description , color=0x00ff00)

    except Exception as e:
        await send_embed(ctx, '', f'Error: {str(e)}', color=0xff0000)

# Botを実行
bot.run(config.DISCORD_TOKEN)
